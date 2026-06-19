#!/usr/bin/env python3
"""Render a canonical test-case model (JSON) to a styled .xlsx workbook.

The xlsx format adapter for the generating-test-cases skill. The case CONTENT
(coverage, evidence) is produced by the skill; this only renders it. Keeping the
adapter separate is the point — the same JSON can be rendered to md/csv/tracker.

Usage:
    python render_xlsx.py cases.json out.xlsx

Input JSON:
{
  "title": "Feature X — Test Cases",
  "meta": {"Branch": "...", "Build": "...", "Automated suite": "..."},   # optional, -> Summary
  "hard_gates": "TC-07, TC-12 (...)",                                     # optional, -> Summary
  "cases": [                                                             # required
    {"ID","Area","Title","Type","Priority","Auto-test","Preconditions",
     "Steps","Test Data","Expected Result","Evidence","Status","Tester","Date","Notes"}
  ],
  "test_data": [{"Field","Value","Notes"}]                               # optional -> Test Data sheet
}

Every case SHOULD carry an "Evidence" value (the skill's Iron Law). Rows whose
Evidence is empty or whose Expected Result contains a guess-word are written with
a red flag fill so the gap is visible.

Requires: openpyxl (pip install openpyxl).
"""
import json
import re
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

COLS = ["ID", "Area", "Title", "Type", "Priority", "Auto-test", "Preconditions",
        "Steps", "Test Data", "Expected Result", "Evidence", "Status",
        "Tester", "Date", "Notes"]
WIDTHS = [10, 15, 30, 10, 8, 26, 24, 46, 20, 46, 26, 12, 12, 12, 20]
STATUS_COL = COLS.index("Status") + 1          # 1-based
EVID_COL = COLS.index("Evidence") + 1
EXP_COL = COLS.index("Expected Result") + 1
GUESS = re.compile(r"\b(probably|assume[d]?|typically|normally|should be|likely|guess)\b", re.I)

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")
FLAG_FILL = PatternFill("solid", fgColor="FDE2E1")   # missing evidence / guess-word
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)


def _style_header(ws, headers):
    ws.append(headers)
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1] if i - 1 < len(WIDTHS) else 18
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 24


def build(model, out_path):
    cases = model.get("cases", [])
    if not cases:
        raise SystemExit("no 'cases' in input JSON")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    _style_header(ws, COLS)

    flagged = 0
    for c in cases:
        row = [c.get(k, "") for k in COLS]
        ws.append(row)
        r = ws.max_row
        for cell in ws[r]:
            cell.alignment = TOP
            cell.border = BORDER
        if r % 2 == 0:
            for cell in ws[r]:
                cell.fill = ALT_FILL
        # Iron-Law visibility: flag missing evidence or guess-words.
        evid = str(c.get("Evidence", "")).strip()
        exp = str(c.get("Expected Result", ""))
        if not evid or GUESS.search(exp) or GUESS.search(str(c.get("Steps", ""))):
            ws.cell(r, EVID_COL).fill = FLAG_FILL
            ws.cell(r, EXP_COL).fill = FLAG_FILL
            flagged += 1
        if str(c.get("Priority", "")).upper() == "P1":
            ws.cell(r, 5).font = Font(bold=True, color="C00000")
        if str(c.get("Type", "")).lower() == "automated":
            ws.cell(r, 6).font = Font(italic=True, color="385723")

    last = ws.max_row
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"

    sl = get_column_letter(STATUS_COL)
    dv = DataValidation(type="list",
                        formula1='"Not Run,Pass,Fail,Blocked,Pass (auto),N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{sl}2:{sl}{last}")

    def cf(val, bg, fg):
        ws.conditional_formatting.add(
            f"{sl}2:{sl}{last}",
            CellIsRule(operator="equal", formula=[f'"{val}"'],
                       fill=PatternFill("solid", fgColor=bg), font=Font(color=fg, bold=True)))
    cf("Pass", "C6EFCE", "006100")
    cf("Pass (auto)", "B7DEE8", "1F4E78")
    cf("Fail", "FFC7CE", "9C0006")
    cf("Blocked", "FFEB9C", "9C6500")
    cf("Not Run", "EDEDED", "808080")
    cf("N/A", "D9D9D9", "595959")

    # ---- Summary ----
    ws2 = wb.create_sheet("Summary")
    total = len(cases)
    auto = sum(1 for c in cases if str(c.get("Type", "")).lower() == "automated")
    p1 = sum(1 for c in cases if str(c.get("Priority", "")).upper() == "P1")
    ws2.append([model.get("title", "Test Cases"), ""])
    ws2.append(["", ""])
    for k, v in (model.get("meta") or {}).items():
        ws2.append([k, v])
    ws2.append(["", ""])
    ws2.append(["Total cases", total])
    ws2.append(["  Automated", auto])
    ws2.append(["  Manual", total - auto])
    ws2.append(["  Priority P1", p1])
    if flagged:
        ws2.append(["  ⚠ Rows flagged (missing evidence / guess-word)", flagged])
    if model.get("hard_gates"):
        ws2.append(["", ""])
        ws2.append(["Hard gates", model["hard_gates"]])
    ws2.append(["", ""])
    ws2.append(["Status legend", "Not Run / Pass / Fail / Blocked / Pass (auto) / N/A"])
    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 84
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
    for r in range(3, ws2.max_row + 1):
        ws2.cell(r, 1).font = Font(bold=True)
        ws2.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Legend ----
    ws3 = wb.create_sheet("Legend")
    legend = [
        ("Legend", ""), ("", ""),
        ("Priority", ""),
        ("P1", "Core path / regression of a shipped behavior / a hard gate — must pass before sign-off."),
        ("P2", "Important functional or negative case — should pass."),
        ("P3", "Edge, cosmetic, or nice-to-have."),
        ("", ""), ("Type", ""),
        ("Automated", "Covered by an automated test; see the Auto-test column. Status = 'Pass (auto)' when green."),
        ("Manual", "Requires running the app / real hardware to verify."),
        ("", ""), ("Status", ""),
        ("Not Run", "Not yet executed."),
        ("Pass", "Manually verified to meet the Expected Result."),
        ("Fail", "Did not meet Expected Result (record details + defect id in Notes)."),
        ("Blocked", "Cannot run (missing hardware/dependency/precondition)."),
        ("Pass (auto)", "Verified by the automated suite."),
        ("N/A", "Not applicable in this environment."),
        ("", ""), ("Evidence", ""),
        ("(required)", "Where the case/expected came from: doc, code file:line, git ref, or a live-system query. A red-flagged row means it is missing or contains a guess-word — fix before sign-off."),
    ]
    for row in legend:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 96
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E78")
    for hdr in (3, 8, 12, 20):
        ws3.cell(hdr, 1).font = Font(bold=True, color="1F4E78")
    for r in range(4, ws3.max_row + 1):
        ws3.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Test Data ----
    td = model.get("test_data") or []
    if td:
        ws4 = wb.create_sheet("Test Data")
        _style_header(ws4, ["Field", "Value", "Notes"])
        for d in td:
            ws4.append([d.get("Field", ""), d.get("Value", ""), d.get("Notes", "")])
            for cell in ws4[ws4.max_row]:
                cell.alignment = TOP
                cell.border = BORDER
        ws4.column_dimensions["A"].width = 24
        ws4.column_dimensions["B"].width = 46
        ws4.column_dimensions["C"].width = 52
        ws4.freeze_panes = "A2"

    wb.save(out_path)
    print(f"wrote {out_path}: {total} cases ({auto} auto / {total-auto} manual), {p1} P1, {flagged} flagged")


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: render_xlsx.py cases.json out.xlsx")
    with open(sys.argv[1], encoding="utf-8") as f:
        model = json.load(f)
    build(model, sys.argv[2])


if __name__ == "__main__":
    main()
