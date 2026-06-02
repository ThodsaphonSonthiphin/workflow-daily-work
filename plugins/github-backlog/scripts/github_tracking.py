"""
github_tracking.py — add traceability columns to a spreadsheet and write created
GitHub issue numbers/URLs back, matched row-by-row by key.

Subcommands:
  add-columns  --source <xlsx|csv> [--key "#"]
      Ensure these columns exist (appended after the last used column, idempotent):
        Issue # | Issue URL | State | Created
  writeback    --source <xlsx|csv> --result <github_backlog_result.json> [--key "#"]
      Match each result item's `key` to the source key column and fill the
      tracking columns. Idempotent: rows that already have an Issue # are left as-is.
"""
import argparse
import json
from datetime import datetime

TRACKING = ["Issue #", "Issue URL", "State", "Created"]


# ---------- xlsx ----------
def _xlsx_headers(ws):
    return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}


def xlsx_add_columns(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = _xlsx_headers(ws)
    missing = [c for c in TRACKING if c not in headers]
    start = ws.max_column + 1
    for i, name in enumerate(missing):
        ws.cell(row=1, column=start + i, value=name)
    wb.save(path)
    cols = _xlsx_headers(ws)
    from openpyxl.utils import get_column_letter
    print("tracking columns: " + ", ".join(f"{n}={get_column_letter(cols[n])}" for n in TRACKING))


def xlsx_writeback(path, result, key_col):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = _xlsx_headers(ws)
    for needed in TRACKING + [key_col]:
        if needed not in headers:
            raise SystemExit(f"missing column '{needed}' — run add-columns first (key='{key_col}')")
    row_of = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=headers[key_col]).value
        if v is not None:
            row_of[str(v)] = r
    return _do_writeback(result, row_of,
                         setter=lambda r, col, val: ws.cell(row=r, column=headers[col], value=val),
                         saver=lambda: wb.save(path))


# ---------- csv ----------
def csv_load(path):
    import csv
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def csv_save(path, rows):
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def csv_add_columns(path):
    rows = csv_load(path)
    if not rows:
        raise SystemExit("empty csv")
    header = rows[0]
    for name in TRACKING:
        if name not in header:
            header.append(name)
            for r in rows[1:]:
                r.append("")
    csv_save(path, rows)
    print("tracking columns ensured: " + ", ".join(TRACKING))


def csv_writeback(path, result, key_col):
    rows = csv_load(path)
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}
    for needed in TRACKING + [key_col]:
        if needed not in idx:
            raise SystemExit(f"missing column '{needed}' — run add-columns first")
    row_of = {str(r[idx[key_col]]): r for r in rows[1:] if len(r) > idx[key_col]}

    def setter(rowobj, col, val):
        rowobj[idx[col]] = val
    return _do_writeback(result, row_of, setter=setter, saver=lambda: csv_save(path, rows))


# ---------- shared ----------
def _do_writeback(result, row_of, setter, saver):
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    written = 0
    for item in result.get("items", []):
        if not item.get("number"):
            continue
        target = row_of.get(str(item["key"]))
        if target is None:
            print(f"  warn: key {item['key']} not found in source")
            continue
        setter(target, "Issue #", item["number"])
        setter(target, "Issue URL", item.get("url", ""))
        setter(target, "State", "open")
        setter(target, "Created", stamp)
        written += 1
        print(f"  key {item['key']} -> #{item['number']}")
    saver()
    tracking = result.get("tracking_issue") or {}
    if tracking.get("url"):
        print(f"tracking issue: {tracking['url']}")
    print(f"wrote {written} issue links back to source")


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    a = sub.add_parser("add-columns")
    a.add_argument("--source", required=True)
    a.add_argument("--key", default="#")
    w = sub.add_parser("writeback")
    w.add_argument("--source", required=True)
    w.add_argument("--result", required=True)
    w.add_argument("--key", default="#")
    args = ap.parse_args()

    is_csv = args.source.lower().endswith((".csv", ".tsv"))
    if args.cmd == "add-columns":
        (csv_add_columns if is_csv else xlsx_add_columns)(args.source)
    else:
        with open(args.result, encoding="utf-8") as f:
            result = json.load(f)
        (csv_writeback if is_csv else xlsx_writeback)(args.source, result, args.key)


if __name__ == "__main__":
    main()
